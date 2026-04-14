"""
W-9 PDF Data Extraction Pipeline - FIXED VERSION
Key fixes:
1. Improved business_name extraction from scanned PDFs
2. Better SSN/EIN regex for OCR noise
3. Improved date extraction from signature section
4. Always uses openpyxl (no xlwings issues)
"""

import hashlib
import re
import os
import shutil
import logging
import unicodedata
from pathlib import Path

import warnings
warnings.filterwarnings("ignore")
os.environ["TOKENIZERS_PARALLELISM"] = "false"

import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler("w9_extraction.log")],
)
logger = logging.getLogger(__name__)

CHECKBOX_LABELS = {
    0: "Individual/Sole Proprietor",
    1: "C Corporation",
    2: "S Corporation",
    3: "Partnership",
    4: "Trust/Estate",
    5: "LLC",
    6: "Other",
}

FIELD_MAP = {
    "f1_01": "name",
    "f1_02": "business_name",
    "f1_03": "llc_classification",
    "f1_04": "other_classification",
    "f1_05": "exempt_payee_code",
    "f1_06": "fatca_code",
    "f1_07": "address",
    "f1_08": "city_state_zip",
    "f1_09": "account_numbers",
    "f1_10": "requester_name_addr",
    "f1_11": "ssn_part1",
    "f1_12": "ssn_part2",
    "f1_13": "ssn_part3",
    "f1_14": "ein_part1",
    "f1_15": "ein_part2",
}

OUTPUT_COLUMNS = [
    "source_file", "name", "business_name", "tax_classification",
    "llc_classification", "address", "city_state_zip", "ssn", "ein",
    "exempt_payee_code", "fatca_code", "account_numbers", "date",
    "extraction_method", "extraction_notes",
]


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bytes):
        try:
            value = value.decode("utf-8")
        except UnicodeDecodeError:
            value = value.decode("latin-1", errors="replace")
    
    text = str(value)

    text = re.sub(r'[_{}\[\]|]', ' ', text) 
    return unicodedata.normalize("NFKC", text.strip())


def _file_md5(file_path: str) -> str:
    hash_md5 = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()


def is_pdf_scanned(pdf_path: str, pdf=None) -> bool:
    try:
        if pdf is not None:
            text = pdf.pages[0].extract_text() or ""
        else:
            with pdfplumber.open(pdf_path) as pdf_obj:
                text = pdf_obj.pages[0].extract_text() or ""
        return len(text.replace(" ", "").replace("\n", "")) < 50
    except Exception:
        return True


def _pdf_has_acroform_fields(pdf) -> bool:
    try:
        for page in pdf.pages:
            for annot in page.annots:
                data = annot.get("data", {})
                if "Tx" in str(data.get("FT", "")) and data.get("V"):
                    return True
                if "Btn" in str(data.get("FT", "")) and data.get("AS"):
                    return True
        return False
    except Exception:
        return False


def _parse_w9_text(text: str, pdf_path: str) -> dict:
    record = {col: "" for col in OUTPUT_COLUMNS}
    record["source_file"] = Path(pdf_path).name
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    logger.debug("Parsed OCR/text lines:")
    for i, line in enumerate(lines):
        logger.debug(f"  {i}: {repr(line)}")

    # Name: line immediately after the "Name of entity" header line
    record["name"] = _extract_name(lines)
    record["business_name"] = _extract_business_name(lines, text)
    record["tax_classification"] = _detect_classification(text, record["name"], lines)
    record["address"] = _extract_after_label(lines, r"^5\s+Address|Address\s*\(number")
    record["city_state_zip"] = _extract_after_label(lines, r"^6\s+City|City,\s*state")
    record["ssn"] = _extract_ssn(text)
    record["ein"] = _extract_ein(text, ssn=record["ssn"])
    record["date"] = _extract_date(text, lines)
    return record


def _extract_name(lines: list) -> str:
    # Strict list of headers to ignore
    SKIP = re.compile(
        r"Form W-9|Request for Taxpayer|Department of the Treasury|Internal Revenue Service"
        r"|entity.?s name on|line 2|An entry is required|sole proprietor"
        r"|Name\s+of\s+entity|1\s+Name|Go to www\.irs\.gov",
        re.IGNORECASE
    )

    for i, line in enumerate(lines):
        line = line.strip()
        # Look specifically for the "1 Name" anchor
        if re.search(r"1\s*Name\s*of\s*entity", line, re.IGNORECASE):
            # Check the following lines for the actual value
            for j in range(i + 1, i + 3):
                if j < len(lines):
                    candidate = lines[j].strip()
                    if candidate and not SKIP.search(candidate) and len(candidate) > 2:
                        return candidate
    return ""

def _extract_after_label(lines: list, pattern: str) -> str:
    skip = [
        r"^(An entry|For a sole|Check the|Enter your|See instructions|Note:|If different|Business name|disregarded)",
    ]
    alt_pattern = pattern
    if "^6" in pattern:
        alt_pattern = r"^6\s+(?:City|Clty|Cty)|(?:City|Clty|Cty),\s*state"
    for i, line in enumerate(lines):
        if re.search(alt_pattern, line, re.IGNORECASE):
            logger.debug(f"Found pattern '{pattern}' in line {i}: {repr(line)}")
            for j in range(i + 1, min(i + 6, len(lines))):
                c = lines[j]
                logger.debug(f"Checking line {j}: {repr(c)}")
                if not any(re.match(p, c, re.IGNORECASE) for p in skip) and len(c) > 2:
                    if not re.search(pattern, c, re.IGNORECASE):
                        logger.debug(f"Returning {repr(c)}")
                        return c
    logger.debug(f"No match for pattern '{pattern}'")
    return ""


def _clean_name_noise(text: str) -> str:
    """Очищает строку от типичного мусора OCR, сохраняя валидный текст."""
    if not text:
        return ""
    # Заменяем _, {, }, [, ], |, и обратные слеши на пробелы
    cleaned = re.sub(r'[_{}\[\]|\\]', ' ', text)
    # Схлопываем множественные пробелы в один
    cleaned = re.sub(r'\s+', ' ', cleaned)
    # Удаляем не-буквенно-цифровые символы в самом начале и конце строки
    cleaned = re.sub(r'^[^a-zA-Z0-9]+|[^a-zA-Z0-9]+$', '', cleaned)
    return cleaned.strip()

def _detect_classification_in_lines(lines: list) -> str:
    checkbox_marker = r'(?:[xX*]|☒|☑|\[[xX* ]\]|\(\s*[xX*]\s*\))'
    candidates = [
        ("Individual/Sole Proprietor", r'Individual/?sole proprietor'),
        ("C Corporation", r'C\s*Corporat\w*|C\s*Corp\.?'),
        ("S Corporation", r'S\s*Corporat\w*|S\s*Corp\.?'),
        ("Partnership", r'Partnership'),
        ("Trust/Estate", r'Trust/?Estate'),
        ("Other", r'Other'),
        ("LLC", r'LLC'),
    ]

    for line in lines:
        for label, pattern in candidates:
            if re.search(pattern, line, re.IGNORECASE) and re.search(checkbox_marker, line):
                return label

    for i, line in enumerate(lines):
        if re.search(checkbox_marker, line):
            for j in range(i, min(i + 3, len(lines))):
                for label, pattern in candidates:
                    if re.search(pattern, lines[j], re.IGNORECASE):
                        return label
    return ""


def _detect_classification(text: str, name: str = "", lines: list = None) -> str:
    if lines is None:
        lines = [l.strip() for l in text.splitlines() if l.strip()]

    classification = _detect_classification_in_lines(lines)
    if classification:
        return classification

    text_lower = (text or "").lower()
    name_lower = (name or "").lower()
    has_form_template = bool(re.search(r'check the appropriate box|line\s*3a|\b3a\b', text_lower, re.IGNORECASE))

    if not has_form_template:
        if re.search(r'\bc\s*(corporat\w*|corp|corp\.)\b', text_lower):
            return "C Corporation"
        if re.search(r'\bs\s*(corporat\w*|corp|corp\.)\b', text_lower):
            return "S Corporation"
        if re.search(r'\bpartnership\b', text_lower):
            return "Partnership"
        if re.search(r'\btrust/?estate\b', text_lower):
            return "Trust/Estate"
        if re.search(r'\bother\b', text_lower):
            return "Other"
        if re.search(r'\bindividual/sole proprietor\b|\bsole proprietor\b|\bindividual\b', text_lower):
            return "Individual/Sole Proprietor"
        if re.search(r'\bllc\b', text_lower):
            if re.search(r'\bc\s*(corporation|orp|orp\.)\b', text_lower):
                return "C Corporation"
            if re.search(r'\bs\s*(corporation|orp|orp\.)\b', text_lower):
                return "S Corporation"
            if re.search(r'\bpartnership\b', text_lower):
                return "Partnership"
            return "LLC"

    if any(x in name_lower for x in ["inc", "corp", "corporation"]):
        return "C Corporation"
    if any(x in name_lower for x in ["llp", "partnership"]):
        return "Partnership"
    if "llc" in name_lower:
        return "LLC"

    return "Individual/Sole Proprietor"

def _extract_business_name(lines: list, text: str) -> str:
    for i, line in enumerate(lines):
        # Anchor for Line 2
        if re.search(r"2\s*Business\s*name", line, re.IGNORECASE):
            # Check the same line first (after the label)
            if "different from above" in line.lower():
                parts = re.split(r"above\.?", line, flags=re.IGNORECASE)
                if len(parts) > 1:
                    candidate = clean_text(parts[1])
                    candidate = _clean_name_noise(candidate) # Применяем очистку
                    if len(candidate) > 2:
                        return candidate
            
            # Check the next 2 lines
            for j in range(i + 1, i + 3):
                if j < len(lines):
                    candidate = clean_text(lines[j])
                    candidate = _clean_name_noise(candidate) # Применяем очистку
                    # Ensure it's not the start of Line 3 or another section header
                    if candidate and not re.search(r"^(3a|Check|3b|Individual/sole proprietor|LLC|Other|Trust/Estate|C\s*Corporation|S\s*Corporation|Partnership)", candidate, re.IGNORECASE):
                        if len(candidate) > 2:
                            return candidate
    return ""

def _ocr_digits_from_crop(crop):
    try:
        import cv2
        import pytesseract
    except ImportError:
        return ""

    gray = cv2.cvtColor(crop, cv2.COLOR_RGB2GRAY)
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    inv = cv2.bitwise_not(thresh)
    proc = cv2.resize(inv, (inv.shape[1] * 2, inv.shape[0] * 2), interpolation=cv2.INTER_CUBIC)
    candidates = []

    configs = [
        '--oem 3 --psm 6 -c tessedit_char_whitelist="0123456789- "',
        '--oem 3 --psm 7 -c tessedit_char_whitelist="0123456789- "',
        '--oem 3 --psm 11 -c tessedit_char_whitelist="0123456789- "',
    ]

    for cfg in configs:
        text = pytesseract.image_to_string(proc, config=cfg)
        digits = ''.join(re.findall(r'\d', text))
        if digits:
            candidates.append((len(digits), digits, text, cfg))

    if candidates:
        candidates.sort(key=lambda x: (-x[0], x[1]))
        best_digits = candidates[0][1]
        logger.debug(f"OCR digits candidates: {[(c[0], c[1], c[3]) for c in candidates]}")
        return best_digits
    return ""




def _extract_ssn_from_pdf_coordinates(page) -> str:
    """Extract SSN by searching extracted words in TIN region using pdfplumber.
    
    Searches right half of page (x >= 50%), middle-lower region (y 35%-55%)
    for digit sequences matching SSN pattern. Avoids OCR on handwritten digits.
    """
    try:
        words = page.extract_words(x_tolerance=3, y_tolerance=3)
        if not words:
            return ""
        
        page_width = page.width
        page_height = page.height
        x_threshold = page_width * 0.50
        y_min = page_height * 0.35
        y_max = page_height * 0.55
        
        # Collect all text in target region
        region_text = []
        for word in words:
            x0 = word.get("x0", 0)
            top = word.get("top", 0)
            text = word.get("text", "")
            
            if x0 >= x_threshold and y_min <= top <= y_max:
                region_text.append(text)
        
        combined = " ".join(region_text)
        logger.debug(f"TIN region text via pdfplumber: {repr(combined)}")
        
        # Look for patterns
        m = re.search(r'\b(\d{3})\s*[-.\s]\s*(\d{2})\s*[-.\s]\s*(\d{4})\b', combined)
        if m:
            result = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
            logger.info(f"SSN found via pdfplumber coordinates: {result}")
            return result
        
        # Try 9 consecutive digits
        for m in re.finditer(r'\b(\d{3})[\s\-]*(\d{2})[\s\-]*(\d{4})\b', combined):
            area = m.group(1)
            if area not in ['000', '666'] and not (900 <= int(area) <= 999):
                result = f"{area}-{m.group(2)}-{m.group(3)}"
                logger.info(f"SSN found via pdfplumber (block): {result}")
                return result
    except Exception as e:
        logger.debug(f"pdfplumber SSN extraction failed: {e}")
    
    return ""


def _extract_ein_from_pdf_coordinates(page) -> str:
    """Extract EIN by searching extracted words in TIN region using pdfplumber.
    
    Searches right half of page (x >= 50%), lower region (y 50%-70%)
    for digit sequences matching EIN pattern. Avoids OCR on handwritten digits.
    """
    try:
        words = page.extract_words(x_tolerance=3, y_tolerance=3)
        if not words:
            return ""
        
        page_width = page.width
        page_height = page.height
        x_threshold = page_width * 0.50
        y_min = page_height * 0.50
        y_max = page_height * 0.70
        
        # Collect all text in target region
        region_text = []
        for word in words:
            x0 = word.get("x0", 0)
            top = word.get("top", 0)
            text = word.get("text", "")
            
            if x0 >= x_threshold and y_min <= top <= y_max:
                region_text.append(text)
        
        combined = " ".join(region_text)
        logger.debug(f"EIN region text via pdfplumber: {repr(combined)}")
        
        # Look for patterns
        m = re.search(r'\b(\d{2})\s*[-.\s]\s*(\d{7})\b', combined)
        if m:
            result = f"{m.group(1)}-{m.group(2)}"
            logger.info(f"EIN found via pdfplumber coordinates: {result}")
            return result
        
        # Try 9 consecutive digits
        for m in re.finditer(r'\b(\d{2})[\s\-]*(\d{7})\b', combined):
            result = f"{m.group(1)}-{m.group(2)}"
            logger.info(f"EIN found via pdfplumber (block): {result}")
            return result
    except Exception as e:
        logger.debug(f"pdfplumber EIN extraction failed: {e}")
    
    return ""


def _extract_ssn_from_image(img) -> str:
    try:
        import cv2
        import numpy as np
    except ImportError:
        return ""

    h, w = img.shape[:2]
    candidate_regions = [
        (int(w * 0.45), int(h * 0.42), w, int(h * 0.52)),
        (int(w * 0.45), int(h * 0.44), w, int(h * 0.54)),
        (int(w * 0.47), int(h * 0.43), w, int(h * 0.56)),
    ]

    for x0, y0, x1, y1 in candidate_regions:
        crop = img[y0:y1, x0:x1]
        if crop.size == 0:
            continue
        digits = _ocr_digits_from_crop(crop)
        logger.debug(f"SSN crop digits: {digits} from region {x0},{y0},{x1},{y1}")
        if len(digits) >= 9:
            return f"{digits[:3]}-{digits[3:5]}-{digits[5:9]}"
    return ""


def _extract_ein_from_image(img) -> str:
    try:
        import cv2
        import numpy as np
    except ImportError:
        return ""

    h, w = img.shape[:2]
    candidate_regions = [
        (int(w * 0.45), int(h * 0.52), w, int(h * 0.63)),
        (int(w * 0.45), int(h * 0.54), w, int(h * 0.66)),
        (int(w * 0.47), int(h * 0.50), w, int(h * 0.68)),
    ]

    for x0, y0, x1, y1 in candidate_regions:
        crop = img[y0:y1, x0:x1]
        if crop.size == 0:
            continue
        digits = _ocr_digits_from_crop(crop)
        logger.debug(f"EIN crop digits: {digits} from region {x0},{y0},{x1},{y1}")
        if len(digits) >= 9:
            return f"{digits[:2]}-{digits[2:9]}"
    return ""


def _extract_ssn(text: str) -> str:
    """Extract SSN in various formats (XXX-XX-XXXX, XXX XX XXXX, XXXXXXXXX, etc)."""
    if not text:
        return ""
    
    # Clean OCR noise but keep structure
    cleaned_text = re.sub(r'[_{}\[\]|\\]', ' ', text)
    
    # Pattern 1: Standard format (XXX-XX-XXXX or XXX XX XXXX or with dots)
    m = re.search(r'\b(\d{3})\s*[-.\s]\s*(\d{2})\s*[-.\s]\s*(\d{4})\b', cleaned_text)
    if m:
        result = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        logger.info(f"SSN found via Pattern 1: {result}")
        return result
    
    # Pattern 2: Look for various SSN labels and capture digits after
    m_label = re.search(
        r"social\s+security\s+number[:\s]*([\d\s\-\.\|]{9,40})",
        cleaned_text,
        re.IGNORECASE,
    )
    if m_label:
        digits = re.sub(r"[^\d]", "", m_label.group(1))
        if len(digits) >= 9:
            result = f"{digits[:3]}-{digits[3:5]}-{digits[5:9]}"
            logger.info(f"SSN found via Pattern 2 (label): {result}")
            return result

    # Pattern 2b: Look for the shorter label 'SSN' as a fallback
    m_ssn_label = re.search(r"\bSSN\b[:\s]*([\d\s\-\.]{9,30})", cleaned_text, re.IGNORECASE)
    if m_ssn_label:
        digits = re.sub(r"[^\d]", "", m_ssn_label.group(1))
        if len(digits) >= 9:
            result = f"{digits[:3]}-{digits[3:5]}-{digits[5:9]}"
            logger.info(f"SSN found via Pattern 2b (SSN): {result}")
            return result
    
    # Pattern 3: Look for 9 consecutive digits (with minimal separators)
    for m in re.finditer(r'(\d{3})\s*[-.\s]*(\d{2})\s*[-.\s]*(\d{4})', cleaned_text):
        area, group, serial = m.group(1), m.group(2), m.group(3)
        # Skip invalid SSNs: area 000, 666, or 900-999
        if area not in ['000', '666'] and not (900 <= int(area) <= 999):
            result = f"{area}-{group}-{serial}"
            logger.info(f"SSN found via Pattern 3: {result}")
            return result
    
    # Pattern 4: Fall back to any 9-digit sequence
    m_fallback = re.search(r'\b(\d{9})\b', cleaned_text)
    if m_fallback:
        digits = m_fallback.group(1)
        result = f"{digits[:3]}-{digits[3:5]}-{digits[5:9]}"
        logger.info(f"SSN found via Pattern 4 (fallback): {result}")
        return result
    
    # Log OCR extraction findings for debugging
    logger.info(f"No SSN found. Text sample: {cleaned_text[:200]}")
    
    return ""

def _extract_ein(text: str, ssn: str = "") -> str:
    """Extract EIN in various formats (XX-XXXXXXX, XX XXXXXXX, XXXXXXXXX, etc)."""
    if not text:
        return ""
    
    cleaned_text = re.sub(r'[_{}\[\]|\\]', ' ', text)
    
    # Pattern 1: Look for "Employer Identification Number" label and capture digits after
    m_label = re.search(r"Employer\s+Identification\s+Number[:\s]+([\d\s\-\.]{7,30})", cleaned_text, re.IGNORECASE)
    if m_label:
        digits = re.sub(r"[^\d]", "", m_label.group(1))
        if len(digits) >= 9:
            return f"{digits[:2]}-{digits[2:9]}"

    m_ein_label = re.search(r"\bEIN\b[:\s]*([\d\s\-\.]{7,30})", cleaned_text, re.IGNORECASE)
    if m_ein_label:
        digits = re.sub(r"[^\d]", "", m_ein_label.group(1))
        if len(digits) >= 9:
            return f"{digits[:2]}-{digits[2:9]}"

    if ssn:
        # Avoid selecting an SSN value as EIN if SSN has already been found.
        return ""

    for m in re.finditer(r'\b(\d{2})\s*[-.\s]*(\d{7})\b', cleaned_text):
        ctx_start, ctx_end = max(0, m.start() - 40), min(len(cleaned_text), m.end() + 40)
        ctx = cleaned_text[ctx_start:ctx_end]
        if re.search(r'Employer\s+Identification\s+Number|\bEIN\b|Employer.+Number', ctx, re.IGNORECASE):
            return f"{m.group(1)}-{m.group(2)}"

    if re.search(r'Employer\s+Identification\s+Number|\bEIN\b', cleaned_text, re.IGNORECASE):
        m_fallback = re.search(r'\b(\d{9})\b', cleaned_text)
        if m_fallback:
            digits = m_fallback.group(1)
            return f"{digits[:2]}-{digits[2:9]}"

    return ""


def _extract_date(text: str, lines: list) -> str:
    # Look for date near signature section
    # Pattern: "Sign" ... date nearby
    for i, line in enumerate(lines):
        if re.search(r"^Sign\s|U\.S\.\s+person|Signature\s+of", line, re.IGNORECASE):
            # Check next few lines for a date
            for j in range(i, min(i + 5, len(lines))):
                m = re.search(r"(\d{1,2}[-/\.]\d{1,2}[-/\.]\d{2,4})", lines[j])
                if m:
                    return m.group(1)

    # General date pattern in text
    m = re.search(r"\b(\d{1,2}[-/]\d{1,2}[-/]\d{4})\b", text)
    if m:
        return m.group(1)
    return ""


def extract_scanned_ocr(pdf_path: str) -> dict:
    record = {col: "" for col in OUTPUT_COLUMNS}
    try:
        from pdf2image import convert_from_path
        import pytesseract
        import cv2
        import numpy as np
    except ImportError as e:
        record["extraction_method"] = "OCR_FAILED"
        record["extraction_notes"] = f"Missing: {e}"
        return record

    try:
        images = convert_from_path(pdf_path, dpi=200, first_page=1, last_page=1)
        img = np.array(images[0])
        IMG_W, IMG_H = images[0].size

        logger.info("OCR processing only first page")
        h_px = img.shape[0]
        top_img = img[:int(h_px * 0.75), :]
        full_text = pytesseract.image_to_string(top_img, config="--oem 3 --psm 6")
        record = _parse_w9_text(full_text, pdf_path)

        # Try to extract SSN/EIN via pdfplumber coordinates first (more accurate for scanned forms)
        try:
            with pdfplumber.open(pdf_path) as pdf:
                page = pdf.pages[0]
                if not record.get("ssn"):
                    ssn_from_coords = _extract_ssn_from_pdf_coordinates(page)
                    if ssn_from_coords:
                        record["ssn"] = ssn_from_coords
                        logger.debug("SSN found via pdfplumber coordinates")
                
                if not record.get("ein"):
                    ein_from_coords = _extract_ein_from_pdf_coordinates(page)
                    if ein_from_coords:
                        record["ein"] = ein_from_coords
                        logger.debug("EIN found via pdfplumber coordinates")
        except Exception as e:
            logger.debug(f"pdfplumber coordinate extraction failed: {e}")

        # fill missing OCR-only fields from the image region if the parsed text did not include them
        if not record.get("ssn"):
            ssn_found = _extract_ssn_from_image(img)
            if ssn_found:
                record["ssn"] = ssn_found
                logger.debug(f"SSN found from image: {ssn_found}")
            else:
                logger.debug("SSN not found from image")
        if not record.get("ein"):
            ein_found = _extract_ein_from_image(img)
            if ein_found:
                record["ein"] = ein_found
                logger.debug(f"EIN found from image: {ein_found}")
            else:
                logger.debug("EIN not found from image")

        # --- Date: crop the signature section at bottom right ---
        # Look in the bottom 20% of the page, right half for date patterns
        date_strip = img[int(h_px * 0.8):, int(IMG_W * 0.5):]
        if date_strip.size > 0:
            dg = cv2.cvtColor(date_strip, cv2.COLOR_RGB2GRAY)
            _, dp = cv2.threshold(dg, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            dp = cv2.resize(dp, (dp.shape[1]*2, dp.shape[0]*2), interpolation=cv2.INTER_CUBIC)
            date_raw = pytesseract.image_to_string(dp, config="--oem 3 --psm 6").strip()
            logger.debug(f"Date crop OCR text: {repr(date_raw)}")
            dm = re.search(r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})", date_raw)
            if dm:
                record["date"] = dm.group(1)

        record['extraction_method'] = 'OCR'
    except Exception as e:
        logger.error(f"OCR failed: {e}")
        record["extraction_notes"] = str(e)
        record["extraction_method"] = "OCR_FAILED"
    return record


def extract_text_based(pdf_path: str) -> dict:
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        text = page.extract_text(layout=True) or ""
    record = _parse_w9_text(text, pdf_path)
    record["extraction_method"] = "TextPDF"
    return record


def extract_w9(pdf_path: str) -> dict:
    record = {col: "" for col in OUTPUT_COLUMNS}
    record["source_file"] = Path(pdf_path).name

    logger.info(f"Processing: {record['source_file']}")

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if _pdf_has_acroform_fields(pdf):
                method = "AcroForm"
                data = _extract_acroform_from_pdf(pdf)

            elif is_pdf_scanned(pdf_path, pdf):
                method = "OCR"
                data = extract_scanned_ocr(pdf_path)

            else:
                method = "TextPDF"
                data = extract_text_based(pdf_path)

        for k, v in data.items():
            if v: 
                record[k] = v

        record["extraction_method"] = data.get("extraction_method", method)
        record["extraction_status"] = "SUCCESS"

    except Exception as e:
        logger.error(f"Failed: {e}")
        record["extraction_notes"] = str(e)
        record["extraction_method"] = "FAILED"
        record["extraction_status"] = "FAILED"

    if not any([
        record.get("name"),
        record.get("ssn"),
        record.get("ein")
    ]):
        record["extraction_notes"] = (record.get("extraction_notes", "") + " | EMPTY_RECORD").strip()
        if record["extraction_status"] != "FAILED":
            record["extraction_status"] = "EMPTY"

    logger.info(
        f"{record['source_file']} | "
        f"method={record.get('extraction_method')} | "
        f"status={record.get('extraction_status')} | "
        f"name={record.get('name')} | "
        f"ssn={record.get('ssn')} | "
        f"ein={record.get('ein')}"
    )

    logger.debug(f"EXTRACTED from {record['source_file']}: ")
    logger.debug(f"  Name: {record.get('name')}")
    logger.debug(f"  Business Name: {record.get('business_name')}")
    logger.debug(f"  Tax Classification: {record.get('tax_classification')}")
    logger.debug(f"  Address: {record.get('address')}")
    logger.debug(f"  City/State/ZIP: {record.get('city_state_zip')}")
    logger.debug(f"  SSN: {record.get('ssn')}")
    logger.debug(f"  EIN: {record.get('ein')}")
    logger.debug(f"  Date: {record.get('date')}")
    logger.debug(f"  Method: {record.get('extraction_method')}")
    logger.debug(f"  Status: {record.get('extraction_status')}")

    return record




def save_to_excel(records: list, output_path: str) -> str:
    df = pd.DataFrame(records, columns=OUTPUT_COLUMNS)
    client_columns = {
        "name": "Name of Entity(1)",
        "business_name": "Business Name(2)",
        "tax_classification": "TOB(3a)",
        "address": "Address(5)",
        "city_state_zip": "City, state, and ZIP code(6)",
        "ssn": "Social Security Number",
        "ein": "Employer Identification Number",
        "date": "Date",
    }
    df_out = df[list(client_columns.keys())].rename(columns=client_columns)
    out = output_path.replace(".xlsm", ".xlsx")
    df_out.to_excel(out, index=False, sheet_name="W-9 Data")

    wb = load_workbook(out)
    ws = wb.active

    hf = PatternFill(fill_type="solid", start_color="1F4E79", end_color="1F4E79")
    for cell in ws[1]:
        cell.fill = hf
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws.row_dimensions[1].height = 35

    lf = PatternFill(fill_type="solid", start_color="EBF0FA", end_color="EBF0FA")
    for row in range(2, ws.max_row + 1):
        for cell in ws[row]:
            if row % 2 == 0:
                cell.fill = lf
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws.row_dimensions[row].height = 18

    col_widths = {"Name of Entity(1)": 30, "Business Name(2)": 30, "TOB(3a)": 28,
                  "Address(5)": 30, "City, state, and ZIP code(6)": 25,
                  "Social Security Number": 22, "Employer Identification Number": 28, "Date": 18}
    for i, col in enumerate(client_columns.values(), 1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 20)

    ws.freeze_panes = "A2"
    wb.save(out)
    logger.info(f"Saved {len(records)} records to {out}")
    return out


def run_pipeline(input_dir="input_pdfs", output_excel="w9-extractor.xlsx",
                 processed_dir="output_pdfs", move_files=True):
    input_path = Path(input_dir)
    if not input_path.exists():
        logger.error(f"Not found: {input_dir}")
        return

    pdf_files = sorted(input_path.glob("*.pdf"))
    if not pdf_files:
        logger.warning(f"No PDFs in: {input_dir}")
        return

    if move_files:
        Path(processed_dir).mkdir(parents=True, exist_ok=True)

    records = []
    seen_hashes = {}
    for pdf_path in pdf_files:
        file_hash = _file_md5(pdf_path)
        if file_hash in seen_hashes:
            logger.warning(f"Duplicate PDF skipped: {pdf_path} same content as {seen_hashes[file_hash]}")
            if move_files:
                duplicate_dest = Path(processed_dir) / f"duplicate_{pdf_path.name}"
                try:
                    shutil.move(pdf_path, duplicate_dest)
                except Exception as e:
                    logger.warning(f"Duplicate move failed: {e}")
            continue

        seen_hashes[file_hash] = pdf_path
        record = extract_w9(str(pdf_path))
        records.append(record)
        if move_files:
            name = record.get("name", "Unknown")
            safe = re.sub(r'[<>:"/\\|?*]', "_", name)[:80] or "Unknown"
            dest = Path(processed_dir) / f"{safe}_{pdf_path.stem}.pdf"
            try:
                shutil.move(pdf_path, dest)
            except Exception as e:
                logger.warning(f"Move failed: {e}")

    if records:
        save_to_excel(records, output_excel)
        print(f"Done. {len(records)} records processed.")


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--input", default="input_pdfs")
    p.add_argument("--output", default="w9-extractor.xlsx")
    p.add_argument("--archive", default="output_pdfs")
    p.add_argument("--no-move", action="store_true")
    args = p.parse_args()
    run_pipeline(args.input, args.output, args.archive, not args.no_move)
